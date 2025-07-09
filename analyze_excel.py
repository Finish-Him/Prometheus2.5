import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np
from openpyxl import load_workbook

# Configurações para visualização
plt.style.use('ggplot')
sns.set(style="whitegrid")

# Diretório para salvar os gráficos
output_dir = '/home/ubuntu/dota_analysis/output'
os.makedirs(output_dir, exist_ok=True)

# Carregar o arquivo Excel
excel_path = '/home/ubuntu/upload/Database Oráculo 3.0.xlsx'
print(f"Analisando o arquivo: {excel_path}")

# Listar todas as abas do arquivo Excel
wb = load_workbook(filename=excel_path, read_only=True)
sheet_names = wb.sheetnames
print(f"Abas disponíveis no arquivo: {sheet_names}")

# Função para analisar uma aba específica
def analyze_sheet(sheet_name):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        print(f"\nAnálise da aba: {sheet_name}")
        print(f"Dimensões: {df.shape}")
        print("Primeiras 5 linhas:")
        print(df.head())
        print("Colunas disponíveis:")
        print(df.columns.tolist())
        return df
    except Exception as e:
        print(f"Erro ao analisar a aba {sheet_name}: {e}")
        return None

# Analisar abas relevantes para handicap de Kills
relevant_sheets = ['Match Detail', 'Pro Match', 'LOB', 'Treinamento da IA', 'Dataset Previsor']
sheet_data = {}

for sheet in relevant_sheets:
    if sheet in sheet_names:
        sheet_data[sheet] = analyze_sheet(sheet)

# Função para salvar resultados da análise
def save_analysis_results(filename, content):
    with open(os.path.join(output_dir, filename), 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Resultados salvos em: {os.path.join(output_dir, filename)}")

# Análise inicial das estruturas de dados
analysis_results = "# Análise Inicial da Base de Dados Oráculo 3.0\n\n"
analysis_results += "## Estrutura das Abas Relevantes\n\n"

for sheet_name, df in sheet_data.items():
    if df is not None:
        analysis_results += f"### {sheet_name}\n"
        analysis_results += f"- Dimensões: {df.shape}\n"
        analysis_results += f"- Colunas: {', '.join(df.columns.tolist())}\n\n"

save_analysis_results("estrutura_inicial.md", analysis_results)

# Análise específica para handicap de Kills
# Vamos procurar por dados relacionados a kills e handicap nas abas relevantes
kill_analysis = "# Análise de Handicap de Kills\n\n"

# Verificar se temos dados de handicap de kills
for sheet_name, df in sheet_data.items():
    if df is not None:
        kill_columns = [col for col in df.columns if 'kill' in str(col).lower()]
        handicap_columns = [col for col in df.columns if 'handicap' in str(col).lower()]
        
        kill_analysis += f"## Dados de Kills e Handicap na aba {sheet_name}\n\n"
        
        if kill_columns:
            kill_analysis += f"### Colunas relacionadas a Kills:\n- {', '.join(kill_columns)}\n\n"
        else:
            kill_analysis += "Nenhuma coluna relacionada a Kills encontrada nesta aba.\n\n"
            
        if handicap_columns:
            kill_analysis += f"### Colunas relacionadas a Handicap:\n- {', '.join(handicap_columns)}\n\n"
        else:
            kill_analysis += "Nenhuma coluna relacionada a Handicap encontrada nesta aba.\n\n"

save_analysis_results("analise_kills_handicap.md", kill_analysis)

print("Análise inicial concluída. Verificando dados para análise de handicap de Kills...")
